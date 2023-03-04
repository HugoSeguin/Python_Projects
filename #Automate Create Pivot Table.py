
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
#Create Pivot Table

df= pd.read_excel('file_name.xlsx')

df = df[['Gender', 'Product line', 'Total']]
print(df)

df.pivot_table()

pivot_table  = df.pivot_table(index ='Gender', columns='Product line', vlaues='Total', aggfunc='sum').round(0)
pivot_table.to_excel('pivot_table.xlsx', 'Report', startrow=4)

#add chart

load_workbook('pivot_table.xlsx')

sheet = wb('Report')

min_column = wb.active.min_colum
max_column = wb.active.max_colum
min_row = wb.active.min_row
max_row= wb.active.max_row


print(min_column)
print(max_column)
print(min_row)
print(max_row)

barchart = BarChart()

data = Refrence(sheet, min_col= min_colum, Max_col = Max_column, Min_row = Min_row, Max_row = Max_row)

categories = categories(sheet, min_col= min_colum, min_col = min_column, Min_row = Min_row + 1, Max_row = Max_row)

barchat.add_data(data)

barchart.set_categories(categories, titles_from_data=True)

sheet.add_chart(barchart, "B12")

barchat.title = "Sales by Product Line"

barchat.style = 5

wb.save('barchart.xlsx')



