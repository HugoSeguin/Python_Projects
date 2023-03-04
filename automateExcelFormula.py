#Automate Formulas

from openpyxl improt load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('barchart.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.activw.max_column
min_row = wb.active.min_row
max_row= wb.active.Max_row

sheet['B8'] = '=SUM(B6:B7)'
sheet['B8'].style= 'Currency'

sheet['C8'] = '=SUM(C6:C7)'
sheet['C8'].style= 'Currency'

for i in range(min_colum+1, max_column+1)
    print(i)
    print(get_column_letter(i))
    sheet[f'{letter}{max_row+1}8']= f'=SUM({letter}{min_row+1}:{letter}{Max_row}'
    sheet[f'{letter}8'].style = 'Currency'


wb.save('Report.Xlsx')

